import pandas as pd
import joblib
from sklearn.metrics import classification_report, mean_squared_error, mean_absolute_error
import argparse
import json
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from preprocessing import preprocess_data


def resolve_params(config, registry):
    """
    Resolve the full parameter set for a model without instantiating it.

    Mirrors the merge logic in train_model.py:
      registry default_params + experiment config overrides + seed injection.

    Used here to record the complete hyperparameter set in the test metrics JSON.
    """
    model_type = config.get('model_type')
    if not model_type:
        raise ValueError("'model_type' key is missing from the experiment config.")

    if model_type not in registry['models']:
        available = list(registry['models'].keys())
        raise ValueError(
            f"Model type '{model_type}' not found in registry. "
            f"Available models: {available}"
        )

    model_entry = registry['models'][model_type]

    resolved_params = {
        **model_entry.get('default_params', {}),
        **config.get('fixed_params', {})
    }

    seed_param = model_entry.get('seed_param', 'random_state')
    seed = model_entry.get('seed', registry.get('default_seed', 42))
    resolved_params[seed_param] = seed

    return resolved_params


def apply_color_coding(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=ws.max_column - 1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    for row in ws.iter_rows(min_row=2, min_col=ws.max_column - 2, max_col=ws.max_column - 2, max_row=ws.max_row):
        for cell in row:
            if cell.value == "Yes":
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    wb.save(excel_file)
    print(f"Color-coding applied and saved to {excel_file}")


def predict(test_data, model_path, config_file, registry_file, metrics_file, output_dir):
    # Load test data
    data = pd.read_csv(test_data)
    X, y = preprocess_data(data)

    # Load config and registry to resolve full hyperparameter set
    with open(config_file) as f:
        config = json.load(f)
    with open(registry_file) as f:
        registry = json.load(f)

    resolved_params = resolve_params(config, registry)

    # Load the trained model
    model = joblib.load(model_path)
    y_pred = model.predict(X)

    # Map predictions to human-readable labels
    target_names = ['Predicted Not Survive', 'Predicted Survive']
    y_pred_labels = np.where(y_pred == 1, 'Yes', 'No')

    # Add predictions back to the original dataframe
    # Only add ground-truth column if labels were present in the CSV
    if y is not None:
        data['Survived?'] = np.where(y == 1, 'Yes', 'No')
    data['Insure?'] = y_pred_labels

    # Generate output filenames dynamically
    base_filename = os.path.basename(test_data)
    file_name, _ = os.path.splitext(base_filename)
    os.makedirs('output/predicted_data', exist_ok=True)
    output_excel = os.path.join('output/predicted_data', f"{file_name}_prediction_output.xlsx")

    data.to_excel(output_excel, index=False)
    print(f"Predictions saved to {output_excel}")

    apply_color_coding(output_excel)

    # Metrics and classification report — only possible when ground-truth labels are available
    if y is not None:
        test_report = classification_report(y, y_pred, target_names=target_names)
        print("Test set report:\n", test_report)

        os.makedirs(output_dir, exist_ok=True)
        report_filename = os.path.join(output_dir, f"{file_name}_classification_report.txt")
        with open(report_filename, "w") as f:
            f.write(test_report)

        mse = mean_squared_error(y, y_pred)
        rmse = float(np.sqrt(mse))
        mae = mean_absolute_error(y, y_pred)

        print(f"Test set MSE: {mse:.4f}, RMSE: {rmse:.4f}, MAE: {mae:.4f}")

        metrics = {
            "report_type": "test",
            "model_type": config['model_type'],
            "search_strategy": config.get('search_strategy'),
            "scoring": config.get('scoring'),
            "hyperparameters": resolved_params,
            "test": {
                "mse": round(float(mse), 6),
                "rmse": round(rmse, 6),
                "mae": round(float(mae), 6),
                "classification_report": test_report
            }
        }
    else:
        print("No ground-truth labels found in test data — skipping accuracy metrics.")
        metrics = {
            "report_type": "prediction_only",
            "model_type": config['model_type'],
            "search_strategy": config.get('search_strategy'),
            "scoring": config.get('scoring'),
            "hyperparameters": resolved_params
        }

    # Write metrics JSON with fully resolved params
    os.makedirs(os.path.dirname(metrics_file), exist_ok=True)
    with open(metrics_file, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"Metrics saved to {metrics_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Predict using a trained model.")
    parser.add_argument('--test', type=str, help="Path to the test dataset.")
    parser.add_argument('--model', type=str, help="Path to the trained model.")
    parser.add_argument('--config', type=str, help="Path to the experiment configuration JSON file.")
    parser.add_argument('--registry', type=str, help="Path to the model registry JSON file.")
    parser.add_argument('--metrics', type=str, help="Path to save the metrics JSON file.")
    parser.add_argument('--output_dir', type=str, help="Directory to save reports and text outputs.")

    args = parser.parse_args()
    predict(args.test, args.model, args.config, args.registry, args.metrics, args.output_dir)
